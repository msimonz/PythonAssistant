import os
import shutil
import locale
from datetime import datetime
from openpyxl import load_workbook
import calendar
import convert_to_pdf
import send_email


def encontrar_fecha(anio_actual):
    origin = "/Users/simonmarquez/Dropbox/Simon/TRABAJO/Cetus/{anio_actual}/Pagos/{mes_actual}".format(mes_actual=mes_actual, anio_actual=anio_actual)
    return origin

def crear_carpeta_cambio_de_mes(destination, cuenta_de_cobro, formato_de_aprobacion):
    if not os.path.exists(destination):
        os.mkdir(destination)
        print(f"✅ Carpeta creada en: {destination}")
        modificar_archivos(cuenta_de_cobro, formato_de_aprobacion, anio_actual)
    else:
        print(f"ℹ️ La carpeta ya existe en: {destination}")
        modificar_archivos(cuenta_de_cobro, formato_de_aprobacion, anio_actual)


def modificar_archivos(cuenta_de_cobro, formato_de_aprobacion, anio):
    print("Modificando archivos")
    cobro = load_workbook(cuenta_de_cobro)
    aprobacion = load_workbook(formato_de_aprobacion)
    cobro_sheet = cobro.active
    aprobacion_sheet = aprobacion.active
    _, dias_mes_actual = calendar.monthrange(anio, mes_actual_num)
    if dias_mes_actual > 30:
        cobro_sheet["C15"] = "Días laborados desde el 01 al 31 de {mes_actual}".format(mes_actual=mes_actual)
        aprobacion_sheet["J7"] = "1 al 31 de {mes_actual}".format(mes_actual=mes_actual)
        numero = cobro_sheet["H3"]
        numero.value += 1
        cobro_sheet["H3"] = numero.value
        numero = aprobacion_sheet["E7"]
        numero.value += 1
        aprobacion_sheet["E7"] = numero.value
        fecha_actual = datetime.today().strftime("%d/%m/%y")
        cobro_sheet["H5"] = fecha_actual
        aprobacion_sheet["A3"] = "Fecha {fecha_actual}".format(fecha_actual=fecha_actual)
        aprobacion_sheet["B9"] = "Pago del Servicio del 01 al 31 de {mes_actual}".format(mes_actual=mes_actual)
    else:
        cobro_sheet["C15"] = "Días laborados desde el 01 al {dias_mes_actual} de {mes_actual}".format(dias_mes_actual=dias_mes_actual, mes_actual=mes_actual)
        aprobacion_sheet["J7"] = "1 al {dias_mes_actual} de {mes_actual}".format(dias_mes_actual=dias_mes_actual, mes_actual=mes_actual)
        numero = cobro_sheet["H3"]
        numero.value += 1
        cobro_sheet["H3"] = numero.value
        numero = aprobacion_sheet["E7"]
        numero.value += 1
        aprobacion_sheet["E7"] = numero.value
        fecha_actual = datetime.today().strftime("%d/%m/%y")
        cobro_sheet["H5"] = fecha_actual
        aprobacion_sheet["A3"] = "Fecha: {fecha_actual}".format(fecha_actual=fecha_actual)
        aprobacion_sheet["B9"] = "Pago del Servicio del 01 al 31 de {mes_actual}".format(mes_actual=mes_actual)
    cobro.save(cuenta_de_cobro)
    aprobacion.save(formato_de_aprobacion)
    convert_to_pdf.excel_to_pdf(cuenta_de_cobro, cobro_pdf)
    convert_to_pdf.excel_to_pdf(formato_de_aprobacion, aprobacion_pdf)


#FECHAS REQUERIDAS
locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
mes_actual_num = datetime.now().month
anio_actual = datetime.now().year
mes_actual = datetime.now().strftime("%B").capitalize()
if mes_actual_num == 1:
    mes_anterior_num = 12
    anio_anterior = anio_actual - 1 
else:
    mes_anterior_num = mes_actual_num - 1
    anio_anterior = anio_actual
mes_anterior = datetime(anio_anterior, mes_anterior_num, 1).strftime("%B").capitalize()

#RUTAS DE ARCHIVOS UTILIZANDO LAS FECHAS OBTENIDAS PARA EL MES EN CURSO Y CREACIÓN DE CARPETA
cuenta_de_cobro = "/Users/simonmarquez/Dropbox/Simon/TRABAJO/Cetus/{anio_actual}/Pagos/CuentadeCobro.xlsx".format(anio_actual=anio_actual)
formato_de_aprobacion = "/Users/simonmarquez/Dropbox/Simon/TRABAJO/Cetus/{anio_actual}/Pagos/FormatodeAprobacion.xlsx".format(anio_actual=anio_actual)
planilla_pdf = "/Users/simonmarquez/Dropbox/Simon/TRABAJO/Cetus/{anio_anterior}/Pagos/{mes_anterior}/Planilla{mes_anterior}.pdf".format(anio_anterior=anio_anterior, mes_anterior=mes_anterior)
origin = encontrar_fecha(anio_actual)
cobro_pdf = os.path.join(origin)
aprobacion_pdf = os.path.join(origin)
crear_carpeta_cambio_de_mes(origin, cuenta_de_cobro, formato_de_aprobacion)

#DATOS NECESARIOS PARA ENVIAR EL CORREO MENSUAL
#sender = "simon.marquezok@gmail.com"
#sender_token = "lpgr szsn ydle fmsd"
#cobro_pdf = cobro_pdf + "/CuentadeCobro.pdf"
#aprobacion_pdf = aprobacion_pdf + "/FormatodeAprobacion.pdf"

#ENVIAR EL CORREO
#files = {}
#files["CuentadeCobro"] = cobro_pdf
#files["Planilla"] = planilla_pdf
#send_email.send_email("contratistas@cetus.com.co", sender, sender_token, mes_actual, files)
#files.clear()
#files["FormatodeAprobacion"] = aprobacion_pdf
#send_email.send_email(recipient, sender, sender_token, mes_actual, files)




